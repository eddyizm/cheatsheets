# update existing excel workbook. 
import openpyxl
file = r'F:\Docs\existingFile.xlsx'
xfile = openpyxl.load_workbook(file)
sheet = xfile['Sheet 1']
sheet['B5'] = '135'
xfile.save(file)


# Open Excel Data
import openpyxl as px
excel_doc = px.load_workbook(r'<PATH_TO_EXCEL_FILE\myfile.xlsx')
sheet_names = excel_doc.sheetnames
sheet_count = len(sheet_names)
# get sheet by name
mowa = excel_doc['MOWA']
multiple_cells = mowa['A1':'B3']
for row in multiple_cells:
    for cell in row:
        pass
        #print (cell.value)

# access all rows
all_rows = tuple(mowa.rows)
# print (all_rows)
all_columns = tuple(mowa.columns)
print (all_columns)

# **************/* 
# open with win32com     
python -m pip install pywin32
# ***************/
import win32com.client as win32
excel = win32.Dispatch('Excel.Application')
wb = excel.Workbooks.Open('<excelFile>')
ws = wb.Worksheets('<SheetName>')
        
# format excel with win32com (pywin32)
# auto fit sheet
ws.Columns.AutoFit()
# color cell and font
ws.Range('A1:AQ1').Interior.ColorIndex = 46
ws.Range('A1:AQ1').Font.ColorIndex = 2

# font excel colors
# https://access-excel.tips/excel-vba-color-code-list/
ws.Range('A1:CS').Interior.ColorIndex = 23
ws.Range('A1:CS').Font.ColorIndex = 2
# bold font
Font.Bold = True

# keep excel from displaying in order to deal with win32com (pywin32)
excel.Visible = False
excel.ScreenUpdating = False
excel.DisplayAlerts = False
excel.EnableEvents = False
##############################################################################

import xlsxwriter
# Create an new Excel file and add a worksheet.
workbook = xlsxwriter.Workbook('demo.xlsx')
worksheet = workbook.add_worksheet()
# Widen the first column to make the text clearer.
worksheet.set_column('A:A', 20)
# Add a bold format to use to highlight cells.
bold = workbook.add_format({'bold': True})
# Write some simple text.
worksheet.write('A1', 'Hello')
# Text with formatting.
worksheet.write('A2', 'World', bold)
# Write some numbers, with row/column notation.
worksheet.write(2, 0, 123)
worksheet.write(3, 0, 123.456)
# Insert an image.
worksheet.insert_image('B5', 'logo.png')
workbook.close()    

# import EXCEL BINARY files using pyxlsb
import pandas as pd
from pyxlsb import open_workbook as open_xlsb

df = []

with open_xlsb('some.xlsb') as wb:
    with wb.get_sheet(1) as sheet:
        for row in sheet.rows():
            df.append([item.v for item in row])

df = pd.DataFrame(df[1:], columns=df[0])

# import EXCEL BINARY files using xlwings
import xlwings as xw

app = xw.App()
book = xw.Book('file.xlsb')
sheet = book.sheets('sheet_name')
df = sheet.range('A1').options(pd.DataFrame, expand='table').value
book.close()
app.kill()